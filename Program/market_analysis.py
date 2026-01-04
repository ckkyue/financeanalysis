# Import
import datetime as dt
from helper_functions import modify_current_date, get_df, get_excel_filename, stock_market
import openpyxl
from openpyxl.styles import Font, PatternFill
import os
from pandas import ExcelWriter as EW
from plot import *
import scipy.stats as stats
from technicals import *
import re
from openpyxl.styles import Font, PatternFill, Border, Side
import time

def classify(rs_ratio, rs_momentum):
    """
    Classify sectors based on RS-Ratio and RS-Momentum values.

    Parameters:
    - rs_ratio (float): The RS-Ratio value.
    - rs_momentum (float): The RS-Momentum value.

    Returns:
    - str: The classification of the sector.
    """
    if rs_ratio >= 100 and rs_momentum >= 100:
        return "Leading"
    elif rs_ratio >= 100:
        return "Weakening"
    elif rs_momentum > 100:
        return "Improving"
    else:
        return "Lagging"

def _get_cell_styles():
    """Define and return cell styles for Excel formatting."""
    return {
        "yellow_fill": PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid"),
        "red_fill": PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid"),
        "orange_fill": PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid"),
        "red_font": Font(color="FF0000"),
        "green_font": Font(color="008000"),
        "red_top_border": Border(top=Side(style="thin", color="FF0000"))
    }

def _get_column_indices(worksheet):
    """Extract column indices from the header row."""
    header_row = [str(cell.value) if cell.value is not None else "" for cell in worksheet[1]]
    
    column_mapping = {
        "Stock": "stock",
        "Beta": "beta", 
        "Close": "close",
        "Volatility 20 Z-Score": "vol20",
        "Volatility 60 Z-Score": "vol60",
        "SMA 20": "sma20",
        "MVP": "mvp",
        "VCP": "vcp",
        "Sector": "sector"
    }
    
    col_idx = {}
    for i, val in enumerate(header_row):
        if val in column_mapping:
            col_idx[column_mapping[val]] = i
        elif val.startswith("Market Cap (B"):
            col_idx["market_cap"] = i
    
    return col_idx

def _apply_price_formatting(cells, styles):
    """Apply red font formatting if Close < SMA 20."""
    stock_cell, close_cell, sma20_cell = cells
    try:
        if (close_cell.value is not None and sma20_cell.value is not None and 
            float(close_cell.value) < float(sma20_cell.value)):
            stock_cell.font = styles["red_font"]
    except (ValueError, TypeError):
        pass

def _apply_beta_formatting(beta_cell, styles):
    """Apply red font formatting for high beta values."""
    try:
        if beta_cell.value is not None and float(beta_cell.value) > 2:
            beta_cell.font = styles["red_font"]
    except (ValueError, TypeError):
        pass

def _apply_volatility_formatting(vol_cells, styles):
    """Apply color formatting for volatility z-score cells."""
    for cell in vol_cells:
        try:
            val = float(cell.value)
            if val > 2:
                cell.font = styles["red_font"]
            elif val < -1:
                cell.font = styles["green_font"]
        except (ValueError, TypeError):
            pass

def _apply_sector_formatting(sector_cell, sector_classifications, styles):
    """Apply sector classification formatting."""
    leading, improving, weakening, lagging = sector_classifications
    sector_val = sector_cell.value
    
    if sector_val in leading or sector_val in improving:
        sector_cell.fill = styles["yellow_fill"]
    elif sector_val in lagging or sector_val in weakening:
        sector_cell.fill = styles["red_fill"]

def _apply_special_formatting(mvp_cell, vcp_cell, styles):
    """Apply MVP and VCP specific formatting."""
    if mvp_cell.value == "MVP":
        mvp_cell.font = styles["green_font"]
    
    if vcp_cell.value is True:
        vcp_cell.fill = styles["orange_fill"]

def _apply_market_cap_border(stock_cell, market_cap_cell, styles, red_border_applied):
    """Apply red top border for first stock with market cap < 10B."""
    if red_border_applied or market_cap_cell is None:
        return red_border_applied
    
    try:
        if float(market_cap_cell.value) < 10:
            stock_cell.border = Border(
                top=styles["red_top_border"].top,
                left=stock_cell.border.left,
                right=stock_cell.border.right,
                bottom=stock_cell.border.bottom
            )
            return True
    except (ValueError, TypeError):
        pass
    
    return red_border_applied

# Helper to plot, build filename and screen if file exists
def _screen_market(start, current_date, index_name, index_dict, period_long, RS, all_stocks, result_folder, sector_classification):
    plot_sector_industry_selected(current_date, index_name, index_dict, RS=RS, all_stocks=all_stocks, save=True)
    excel_filename = get_excel_filename(current_date, index_name, index_dict, period_long, RS, all_stocks, result_folder)
    if os.path.exists(excel_filename):
        screen_excel(excel_filename, sector_classification)
    else:
        print(f"Excel file not found, skipping {index_name} screen: {excel_filename}")

def screen_excel(excel_filename, sector_excel_classification):
    """
    Screen stocks from an Excel file and apply formatting based on specified criteria.

    Parameters:
    - excel_filename (str): Path to the Excel file containing stock data.
    - sector_excel_classification (dict): Dictionary mapping sector classifications to sector names.
    """
    # Get classified sectors
    sector_sets = {
        key: set(sector_excel_classification.get(key, []))
        for key in ["Leading", "Improving", "Weakening", "Lagging"]
    }
    
    # Load workbook and get styles
    wb = openpyxl.load_workbook(excel_filename)
    ws = wb.active
    styles = _get_cell_styles()
    col_idx = _get_column_indices(ws)
    
    # Track red border application
    red_border_applied = False
    
    # Process each data row
    for row in ws.iter_rows(min_row=2):
        # Extract cells using column indices
        cells = {key: row[idx] for key, idx in col_idx.items()}
        
        # Apply various formatting rules
        _apply_price_formatting(
            (cells["stock"], cells["close"], cells["sma20"]), styles
        )

        _apply_beta_formatting(cells["beta"], styles)

        _apply_volatility_formatting(
            [cells["vol20"], cells["vol60"]], styles
        )
        
        _apply_sector_formatting(
            cells["sector"],
            (sector_sets["Leading"], sector_sets["Improving"],
             sector_sets["Weakening"], sector_sets["Lagging"]),
            styles
        )

        _apply_special_formatting(cells["mvp"], cells["vcp"], styles)
        
        red_border_applied = _apply_market_cap_border(
            cells["stock"], cells.get("market_cap"), styles, red_border_applied
        )
    
    wb.save(excel_filename)
    print(f"Changes made to the Excel file {excel_filename}.")

# Main function
def main():
    # Start of the program
    start = dt.datetime.now()

    # Variables
    all_stocks = True
    period_short = 60
    period_long = 252
    RS = 90
    factors = [0.1, 0.55, 0.35]
    backtest = False

    # Index
    index_name = "^GSPC"
    index_names = ["^HSI", "^GSPC", "^IXIC", "^DJI", "IWM", "FFTY", "QQQE"]
    index_dict = {"^HSI": "HKEX", "^GSPC": "S&P 500", "^IXIC": "NASDAQ Composite", "^DJI": "Dow Jones Industrial Average", 
                  "IWM": "iShares Russell 2000 ETF", "FFTY": "Innovator IBD 50 ETF", "QQQE": "NASDAQ-100 Equal Weighted ETF", 
                  "KWEB": "KraneShares CSI China Internet ETF", "GC=F": "Gold"}
    
    # Modify the current date
    current_date = modify_current_date(start, index_name)

    # US Sectors
    us_sectors = ["XLC", "XLY", "XLP", "XLE", "XLF", "XLV", 
            "XLI", "XLB", "XLRE", "XLK", "XLU"]
    us_sector_dict = {"XLC": "Communication Services", "XLY": "Consumer Discretionary", "XLP": "Consumer Staples", 
                "XLE": "Energy", "XLF": "Financials", "XLV": "Health Care", 
                "XLI": "Industrials", "XLB": "Materials", "XLRE": "Real Estate", 
                "XLK": "Technology", "XLU": "Utilities"}
    us_sector_excel_dict = {"XLC": "Communication Services", "XLY": "Consumer Cyclical", "XLP": "Consumer Defensive", 
                "XLE": "Energy", "XLF": "Financial Services", "XLV": "Healthcare", 
                "XLI": "Industrials", "XLB": "Basic Materials", "XLRE": "Real Estate", 
                "XLK": "Technology", "XLU": "Utilities"}
    
    # HK Sectors
    # Setup dictionaries for Hong Kong sector information based on US sector mappings
    hk_sectors = []
    hk_sector_dict = {}
    hk_sector_excel_dict = {}
    hk_sector_stocks = {}

    # Load the HSI sectors data from CSV file
    hsi_sectors_df = pd.read_csv("Program/hsi_sectors.csv")

    # Build HK sector info and stock lists
    for us_sector, us_sector_excel_name in us_sector_excel_dict.items():
        hk_sector = "HSI" + us_sector[2:]
        hk_sectors.append(hk_sector)
        hk_sector_dict[hk_sector] = us_sector_dict[us_sector]
        hk_sector_excel_dict[hk_sector] = us_sector_excel_name
        hk_sector_stocks[hk_sector] = hsi_sectors_df[hsi_sectors_df["Sector"] == us_sector_excel_name]["Stock"].tolist()

    # Aggregate and save HK sector price data
    for hk_sector, stocks in hk_sector_stocks.items():
        sector_df = pd.DataFrame()
        for stock in stocks:
            stock_df = get_df(stock, current_date)
            stock_df["Percent Change"] = stock_df["Close"].pct_change()
            sector_df[f"Percent Change {stock}"] = stock_df["Percent Change"]
        if not sector_df.empty:
            sector_df["Percent Change"] = sector_df.mean(axis=1)
            sector_df = sector_df.dropna()
            sector_df["Close"] = 100 * (1 + sector_df["Percent Change"]).cumprod()
            sector_df.to_csv(f"Price data/{hk_sector}_{current_date}.csv")

    result_folder = "Result"
    sp500_df = get_df("^GSPC", current_date)
    hsi_df = get_df("^HSI", current_date)

    # Plot all tickers (indices, US and HK sectors)
    plot_all = dt.datetime.now().weekday() == 5 or 6  # Plot all on weekends
    if plot_all:
        for ticker in index_names + us_sectors + hk_sectors:
            df = get_df(ticker, current_date)
            plot_close(ticker, df, MVP_VCP=False, save=True)

    # US sector rotation analysis and RRG plot
    plot_us_rrg = True
    if plot_us_rrg:
        sp500_df = get_JdK(index_names + us_sectors + hk_sectors, sp500_df, current_date)
        us_sector_classification = {k: [] for k in ["Leading", "Weakening", "Improving", "Lagging"]}
        us_sector_excel_classification = {k: [] for k in ["Leading", "Weakening", "Improving", "Lagging"]}
        for sector in us_sectors:
            rs_ratio = sp500_df[f"{sector} JdK RS-Ratio"].iloc[-1]
            rs_momentum = sp500_df[f"{sector} JdK RS-Momentum"].iloc[-1]
            category = classify(rs_ratio, rs_momentum)
            us_sector_classification[category].append(us_sector_dict[sector])
            us_sector_excel_classification[category].append(us_sector_excel_dict[sector])
        for category, sectors_list in us_sector_classification.items():
            print(f"{category} US sectors: {', '.join(sectors_list)}")
        plot_rrg(us_sectors, us_sector_dict, sp500_df, "US", "sector", save=True)

    # HK sector rotation analysis and RRG plot
    plot_hk_rrg = True
    if plot_hk_rrg:
        hsi_df = get_JdK(hk_sectors, hsi_df, current_date)
        hk_sector_classification = {k: [] for k in ["Leading", "Weakening", "Improving", "Lagging"]}
        hk_sector_excel_classification = {k: [] for k in ["Leading", "Weakening", "Improving", "Lagging"]}
        for sector in hk_sectors:
            rs_ratio = hsi_df[f"{sector} JdK RS-Ratio"].iloc[-1]
            rs_momentum = hsi_df[f"{sector} JdK RS-Momentum"].iloc[-1]
            category = classify(rs_ratio, rs_momentum)
            hk_sector_classification[category].append(hk_sector_dict[sector])
            hk_sector_excel_classification[category].append(hk_sector_excel_dict[sector])
        for category, sectors_list in hk_sector_classification.items():
            print(f"{category} HK sectors: {', '.join(sectors_list)}")
        plot_rrg(hk_sectors, hk_sector_dict, hsi_df, "HK", "sector", save=True)

    # Plot JdK indicator for each US sector
    if plot_all:
        for sector in us_sectors:
            plot_JdK(sector, us_sector_dict, sp500_df, save=True)

    # Screen US stocks and apply Excel formatting
    screen_us = True
    if screen_us:
        _screen_market(start, current_date, "^GSPC", index_dict, period_long, 90, all_stocks, result_folder, us_sector_excel_classification)

    # Screen HK stocks and apply Excel formatting
    screen_hk = True
    if screen_hk:
        _screen_market(start, current_date, "^HSI", index_dict, period_long, 80, all_stocks, result_folder, hk_sector_excel_classification)

    # Market breadth analysis and plots for S&P 500
    current_date = modify_current_date(start, "^GSPC")
    sp500_df = get_df("^GSPC", current_date, redownload=True)
    tickers = stock_market(current_date, current_date, "^GSPC", False)
    sp500_df = market_breadth(current_date, sp500_df, tickers)
    sp500_filename = f"Price data/GSPC_{current_date}.csv"
    sp500_df.to_csv(sp500_filename)
    plot_market_breadth("^GSPC", sp500_df, tickers, save=True)
    plot_close("^GSPC", sp500_df, MVP_VCP=False)
    plot_MFI_RSI("^GSPC", sp500_df, save=True)

    # VIX analysis and plot
    plot_vix = True
    if plot_vix:
        vix_df = get_df("^VIX", current_date)
        vix_current_high = round(vix_df["High"].iloc[-1], 2)
        vix_df["SMA 5"] = SMA(vix_df, 5)
        vix_sma5 = vix_df["SMA 5"].iloc[-1]
        if vix_current_high < 26 and vix_sma5 <= 0:
            vix_colour = "Green"
        elif 26 < vix_current_high < 30 or vix_sma5 > 0:
            vix_colour = "Yellow"
        elif vix_current_high > 30:
            vix_colour = "Red"
        else:
            vix_colour = "Unknown"
        plot_close("^VIX", vix_df, save=True)
        print(f"Current VIX: {vix_current_high} ({vix_colour})")

    end = dt.datetime.now()
    print(end, "\n")
    print("The program used", end - start)

# Run the main function
if __name__ == "__main__":
    main()